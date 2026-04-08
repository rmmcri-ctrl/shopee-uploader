import streamlit as st
import pandas as pd
import requests
import base64
import re
from io import BytesIO
from openpyxl import load_workbook
import unicodedata

st.set_page_config(page_title="Shopee Uploader", layout="wide")

# URL JÁ PREENCHIDA COM A SUA
URL_API_SHEET = "https://script.google.com/macros/s/AKfycbyehKcqAheEIjgG8s64WwNpeFblcEsjqSeh4ujViS3AKsh4SxOYRpoH226BcPiQtlqt/exec"
IMGBB_API_KEY = "5ebf9740e61741d80a644637d5602009"

def normalizar_texto(texto):
    texto = str(texto).lower()
    texto = ''.join(c for c in unicodedata.normalize('NFD', texto) if unicodedata.category(c)!='Mn')
    texto = re.sub(r'[^a-z0-9\s]', ' ', texto)
    return texto

@st.cache_data(ttl=5)
def carregar_produtos_da_planilha():
    try:
        resp = requests.get(URL_API_SHEET, timeout=10)
        data = resp.json()
        df = pd.DataFrame(data)
        df = df[(df['status'] == 'pendente') & (df['nome']!='')]
        df['preco'] = pd.to_numeric(df['preco'], errors='coerce').fillna(0)
        return df
    except Exception as e:
        st.error(f"Erro ao carregar produtos da planilha: {e}")
        return pd.DataFrame()

@st.cache_data
def carregar_categorias():
    try:
        df_cat = pd.read_excel("template_shopee_Categoria.xlsx").fillna('')
        col_id = [col for col in df_cat.columns if 'id' in col.lower() and 'categoria' in col.lower()][0]
        col_nome_cats = [col for col in df_cat.columns if 'categoria' in col.lower() and 'id' not in col.lower()]
        dicionario = []
        for _, row in df_cat.iterrows():
            partes_nome = [str(row[col]) for col in col_nome_cats if row[col]]
            texto_completo = " > ".join(partes_nome)
            palavras_chave = normalizar_texto(texto_completo)
            id_categoria = row[col_id]
            if id_categoria and texto_completo:
                dicionario.append({"id": str(int(id_categoria)), "nome_completo": texto_completo.strip(), "palavras": palavras_chave})
        return dicionario
    except: return []

def rehost_imgbb(url_original):
    if not url_original or "i.ibb.co" in str(url_original): return url_original
    try:
        resp = requests.get(url_original, headers={'User-Agent': 'Mozilla/5.0'}, timeout=15)
        if resp.status_code == 200:
            img_b64 = base64.b64encode(resp.content)
            payload = {"key": IMGBB_API_KEY, "image": img_b64}
            upload = requests.post("https://api.imgbb.com/1/upload", data=payload, timeout=30).json()
            if upload.get("success"): return upload["data"]["url"]
    except: pass
    return None

def gerar_excel_shopee(df_produtos):
    if df_produtos.empty: return None
    wb = load_workbook("template_shopee.xlsx")
    ws = wb.active
    header_map = {str(cell.value).replace("\n", " ").strip(): col_idx for col_idx, cell in enumerate(ws[3], 1) if cell.value}

    for i, row in df_produtos.iterrows():
        linha = 7 + i
        url_imgbb = rehost_imgbb(row['url_imagem'])
        urls = [f"{url_imgbb}?v={j}" for j in range(1, 4)] if url_imgbb else [None]*3
        dados = {
            "Nome do Produto": str(row['nome'])[:120], "Preço": float(row['preco']),
            "Imagem de Capa": urls[0], "Imagem 1": urls[0], "Imagem 2": urls[1], "Imagem 3": urls[2],
            "Categoria": row.get('categoria_id', ''), "Estoque": 10, "Peso": 0.5,
            "Comprimento": 10, "Largura": 10, "Altura": 10,
            "Descrição do Produto": f"Produto: {row['nome']}<br>Envio rápido. Garantia 7 dias.",
            "Marca": "Sem marca", "Condição": "Novo", "Material": "Plástico",
            "País de Origem": "China", "Duração da Garantia": "7 dias"
        }
        for col_nome, valor in dados.items():
            if col_nome in header_map and valor:
                ws.cell(row=linha, column=header_map[col_nome], value=valor)

    output = BytesIO()
    wb.save(output)
    return output.getvalue()

DIC_CATEGORIAS = carregar_categorias()
st.title("🚀 Painel Shopee Uploader")
st.caption("Conectado com a extensão GDrop via Google Sheets")

df_produtos = carregar_produtos_da_planilha()

tab1, tab2 = st.tabs(["📋 Gerenciar Produtos da Extensão", "📥 Exportar Excel"])

with tab1:
    st.subheader(f"Produtos recebidos da extensão: {len(df_produtos)}")
    st.caption("Atualiza automático a cada 5 segundos. Se não aparecer, clica em Rerun no menu ⋮")

    if not df_produtos.empty:
        for idx, row in df_produtos.iterrows():
            with st.container(border=True):
                col1, col2 = st.columns([1, 3])
                with col1:
                    st.image(row['url_imagem'], width=150)
                with col2:
                    st.markdown(f"**{row['nome']}**")
                    st.markdown(f"**Preço:** R$ {float(row['preco']):.2f}")

                    busca_cat = st.text_input("Buscar categoria", key=f"busca_{idx}", placeholder="Digite: fone, blusa...")
                    cats_filtradas = DIC_CATEGORIAS
                    if busca_cat:
                        busca_norm = normalizar_texto(busca_cat)
                        cats_filtradas = [c for c in DIC_CATEGORIAS if busca_norm in c["palavras"]]

                    opcoes_cat = {c["nome_completo"]: c["id"] for c in cats_filtradas[:50]}
                    opcoes_cat["-- Nenhuma / Preencher depois --"] = ""
                    cat_selecionada_nome = st.selectbox("Selecione a Categoria", options=list(opcoes_cat.keys()), key=f"cat_{idx}")
                    df_produtos.at[idx, 'categoria_id'] = opcoes_cat[cat_selecionada_nome]
                    df_produtos.at[idx, 'categoria_nome'] = cat_selecionada_nome

                    if opcoes_cat[cat_selecionada_nome]:
                        st.success(f"ID da Categoria: {opcoes_cat[cat_selecionada_nome]}")
    else:
        st.info("Nenhum produto recebido ainda. Use a extensão GDrop na Shopee/AliExpress e clique em 'Adicionar ao Dashboard'.")

with tab2:
    st.subheader("Exportar para Shopee")
    if not df_produtos.empty and 'categoria_id' in df_produtos.columns:
        df_final = df_produtos[df_produtos['categoria_id']!='']
        st.markdown(f"**{len(df_final)} produtos** com categoria definida prontos para exportar.")
        if not df_final.empty:
            excel_bytes = gerar_excel_shopee(df_final)
            st.download_button("📥 Baixar Excel Pronto pra Shopee", data=excel_bytes, file_name="shopee_upload.xlsx", type="primary")
        else:
            st.warning("Defina a categoria dos produtos na aba 'Gerenciar' antes de exportar.")
    else:
        st.info("Adicione produtos pela extensão e defina as categorias.")
